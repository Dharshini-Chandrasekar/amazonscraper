from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')
options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

search_url = 'https://www.amazon.com/s?k=laptop'

driver.get(search_url)
time.sleep(3) 

page_source = driver.page_source

soup = BeautifulSoup(page_source, 'html.parser')

products = []
for product in soup.find_all('div', {'data-component-type': 's-search-result'}):
    try:
        name = product.h2.text.strip()
        price = product.find('span', 'a-price-whole').text.strip() if product.find('span', 'a-price-whole') else 'N/A'
        rating = product.find('span', 'a-icon-alt').text.strip() if product.find('span', 'a-icon-alt') else 'N/A'
        reviews = product.find('span', {'class': 'a-size-base'}).text.strip() if product.find('span', {'class': 'a-size-base'}) else 'N/A'
        availability = 'Available' if product.find('span', {'class': 'a-size-small'}) else 'Out of Stock'
        product_link = 'https://www.amazon.com' + product.h2.a['href']

        products.append({
            'Product Name': name,
            'Price': price,
            'Rating': rating,
            'Reviews': reviews,
            'Availability': availability,
            'Product Link': product_link
        })
    except Exception as e:
        continue

wb = Workbook()
ws = wb.active

ws.append(['Product Name', 'Price', 'Rating', 'Reviews', 'Availability', 'Product Link'])

for index, item in enumerate(products, start=2):
    ws[f'A{index}'] = item['Product Name']
    ws[f'B{index}'] = item['Price']

    try:
        rating_value = float(item['Rating'].split()[0])
        stars = '★' * int(rating_value) + '☆' * (5 - int(rating_value))
        ws[f'C{index}'] = stars
    except:
        ws[f'C{index}'] = 'N/A'

    ws[f'D{index}'] = item['Reviews']
    ws[f'E{index}'] = item['Availability']

    ws[f'F{index}'].value = 'Click here'
    ws[f'F{index}'].hyperlink = item['Product Link']
    ws[f'F{index}'].font = Font(color="0000FF", underline="single")  

for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 25

wb.save('amazon_products_with_links.xlsx')

driver.quit()

print("Scraping and saving complete!")
